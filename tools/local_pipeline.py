#!/usr/bin/env python3
"""Generate preprocessed dashboard data from local Excel files and push updates.

This script reads Excel workbooks from local disk, converts a specific worksheet
into the JSON format expected by the dashboard, writes the output into the
preprocessed/ folder, and commits + pushes the changes to this repository.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import subprocess
from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import load_workbook

DEFAULT_CONFIG = Path(__file__).resolve().parent / "local_pipeline_config.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG,
        help="Path to the JSON configuration file (default: tools/local_pipeline_config.json)",
    )
    parser.add_argument(
        "--only",
        action="append",
        default=[],
        help="Process only the dataset with this name (can be provided multiple times)",
    )
    parser.add_argument(
        "--skip-push",
        action="store_true",
        help="Commit changes but skip git push",
    )
    parser.add_argument(
        "--skip-git",
        action="store_true",
        help="Generate files but skip git add/commit/push",
    )
    parser.add_argument(
        "--commit-message",
        default=None,
        help="Override the commit message used when publishing updates",
    )
    return parser.parse_args()


def to_python_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        if value.tzinfo is not None:
            value = value.astimezone(dt.timezone.utc)
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def trim_trailing(values: List[object]) -> List[object]:
    idx = len(values)
    while idx and values[idx - 1] is None:
        idx -= 1
    return values[:idx]


def normalize_row(values: Iterable[object], width: int) -> List[object]:
    cleaned = [to_python_value(v) for v in values]
    cleaned = trim_trailing(cleaned)
    cleaned = cleaned[:width]
    if len(cleaned) < width:
        cleaned.extend([None] * (width - len(cleaned)))
    return cleaned


def write_json_rows(
    output_path: Path,
    rows: Iterable[List[object]],
    source: str,
    sheet_name: str,
    header_row_index: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as fp:
        fp.write("{\n")
        fp.write(f"  \"source\": {json.dumps(source, ensure_ascii=False)},\n")
        fp.write(f"  \"sheet_name\": {json.dumps(sheet_name, ensure_ascii=False)},\n")
        fp.write(f"  \"header_row_index\": {header_row_index},\n")
        generated_at = dt.datetime.utcnow().replace(tzinfo=dt.timezone.utc).isoformat()
        fp.write(f"  \"generated_at\": {json.dumps(generated_at)},\n")
        fp.write("  \"sheet_data\": [\n")
        first = True
        for row in rows:
            if not first:
                fp.write(",\n")
            fp.write("    ")
            fp.write(json.dumps(row, ensure_ascii=False))
            first = False
        fp.write("\n  ]\n}\n")


def iter_sheet_rows(
    excel_path: Path,
    sheet_name: str,
    header_row_index: int,
) -> Iterable[List[object]]:
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {excel_path.name}. Available: {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        pre_rows: List[List[object]] = []
        target_width: Optional[int] = None
        yielded_header = False
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            values = [to_python_value(v) for v in row]
            values = trim_trailing(values)
            if idx < header_row_index:
                pre_rows.append(values)
                continue
            if idx == header_row_index:
                if not values:
                    raise ValueError("Header row appears to be empty")
                target_width = len(values)
                for pre in pre_rows:
                    yield normalize_row(pre, target_width)
                yield normalize_row(values, target_width)
                yielded_header = True
                continue
            if target_width is None:
                raise ValueError("Header row not found while iterating the worksheet")
            yield normalize_row(values, target_width)
        if not yielded_header:
            raise ValueError("No usable sheet found (header row missing)")
    finally:
        workbook.close()


def resolve_repo_root() -> Path:
    root = subprocess.check_output(["git", "rev-parse", "--show-toplevel"], text=True).strip()
    return Path(root)


def load_config(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")
    with path.open("r", encoding="utf-8") as fp:
        return json.load(fp)


def process_workbooks(config: dict, only_names: List[str]) -> List[Path]:
    repo_root = resolve_repo_root()
    outputs: List[Path] = []
    workbooks = config.get("workbooks", [])
    if not isinstance(workbooks, list) or not workbooks:
        raise ValueError("Config must include a non-empty 'workbooks' list")

    for entry in workbooks:
        name = str(entry.get("name", "")).strip()
        if only_names and name not in only_names:
            continue
        excel_path_raw = str(entry.get("excel_path", "")).strip()
        sheet_name = str(entry.get("sheet_name", "")).strip()
        output_path_raw = str(entry.get("output_path", "")).strip()
        header_row_index = int(entry.get("header_row_index", 1))
        if not excel_path_raw:
            raise ValueError(f"Workbook entry '{name or output_path_raw}' is missing excel_path")
        if not sheet_name:
            raise ValueError(f"Workbook entry '{name or output_path_raw}' is missing sheet_name")
        if not output_path_raw:
            raise ValueError(f"Workbook entry '{name or excel_path_raw}' is missing output_path")

        excel_path = Path(excel_path_raw).expanduser().resolve()
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        output_path = Path(output_path_raw)
        if not output_path.is_absolute():
            output_path = repo_root / output_path

        rows = iter_sheet_rows(excel_path, sheet_name, header_row_index)
        write_json_rows(
            output_path=output_path,
            rows=rows,
            source=excel_path.name,
            sheet_name=sheet_name,
            header_row_index=header_row_index,
        )
        print(f"Wrote {output_path}")
        outputs.append(output_path)

    if only_names and not outputs:
        raise ValueError(f"No matching datasets found for --only {only_names}")
    return outputs


def git_publish(paths: List[Path], commit_message: str, push: bool) -> None:
    if not paths:
        print("No output files were generated; skipping git publish.")
        return
    rel_paths = [str(path) for path in paths]
    subprocess.run(["git", "add", "--"] + rel_paths, check=True)
    status = subprocess.check_output(
        ["git", "status", "--porcelain", "--"] + rel_paths,
        text=True,
    ).strip()
    if not status:
        print("No changes detected in generated files; nothing to commit.")
        return
    subprocess.run(["git", "commit", "-m", commit_message], check=True)
    if push:
        subprocess.run(["git", "push"], check=True)


def main() -> None:
    args = parse_args()
    config = load_config(args.config)
    outputs = process_workbooks(config, args.only)
    if args.skip_git:
        print("Skipping git add/commit/push as requested.")
        return
    commit_message = args.commit_message or config.get(
        "commit_message", "Update preprocessed dashboard data"
    )
    git_publish(outputs, str(commit_message), push=not args.skip_push)


if __name__ == "__main__":
    main()
